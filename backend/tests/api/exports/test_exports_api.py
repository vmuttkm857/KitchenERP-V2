from io import BytesIO
from openpyxl import load_workbook
from pypdf import PdfReader
from sqlalchemy import event
from app.db.session import engine as process_engine
from tests.api.kitchen_operations.test_kitchen_operations_api import kitchen_fixture
from tests.api.menus.test_menus_api import foundations,full_structure,meals,menu
from tests.api.requirements.test_requirements_api import auth

def test_exports_require_auth(client):
    response=client.post("/api/v1/exports/kitchen-operations/xlsx",json={"menu_id":"00000000-0000-0000-0000-000000000001"})
    assert response.status_code==401

def test_kitchen_excel_and_pdf_are_binary_authenticated_downloads(client,db_session):
    headers=auth(client,db_session);menu,*_=kitchen_fixture(client,headers,db_session);body={"menu_id":menu["id"],"selected_dates":["2026-10-01"]}
    xlsx=client.post("/api/v1/exports/kitchen-operations/xlsx",headers=headers,json=body)
    assert xlsx.status_code==200,xlsx.text;assert xlsx.headers["content-type"].startswith("application/vnd.openxmlformats")
    assert "attachment" in xlsx.headers["content-disposition"];assert xlsx.content[:2]==b"PK"
    wb=load_workbook(BytesIO(xlsx.content));assert "備料明細" in wb.sheetnames and wb["備料明細"].max_row>1
    pdf=client.post("/api/v1/exports/kitchen-operations/pdf",headers=headers,json=body)
    assert pdf.status_code==200,pdf.text;assert pdf.headers["content-type"].startswith("application/pdf");assert pdf.content.startswith(b"%PDF")
    assert len(PdfReader(BytesIO(pdf.content)).pages)>=1

def test_kitchen_a4_excel_is_an_independent_print_download(client,db_session):
    headers=auth(client,db_session);menu,*_=kitchen_fixture(client,headers,db_session);body={"menu_id":menu["id"],"selected_dates":["2026-10-01"]}
    response=client.post("/api/v1/exports/kitchen-operations/a4-xlsx",headers=headers,json=body)
    assert response.status_code==200,response.text
    assert "A4" in response.headers["content-disposition"]
    wb=load_workbook(BytesIO(response.content))
    assert wb.sheetnames
    assert all(ws.page_setup.paperSize==9 and ws.page_setup.fitToWidth==1 for ws in wb.worksheets)

def test_kitchen_a4_export_query_count_does_not_grow_per_row(client,db_session):
    headers=auth(client,db_session);menu,*_=kitchen_fixture(client,headers,db_session);statements=[]
    def record(conn,cursor,statement,parameters,context,executemany):statements.append(statement)
    event.listen(process_engine,"before_cursor_execute",record)
    try:response=client.post("/api/v1/exports/kitchen-operations/a4-xlsx",headers=headers,json={"menu_id":menu["id"]})
    finally:event.remove(process_engine,"before_cursor_execute",record)
    assert response.status_code==200
    assert len([sql for sql in statements if sql.lstrip().upper().startswith("SELECT")])<=4

def test_kitchen_export_query_count_does_not_grow_per_row(client,db_session):
    headers=auth(client,db_session);menu,*_=kitchen_fixture(client,headers,db_session);statements=[]
    def record(conn,cursor,statement,parameters,context,executemany):statements.append(statement)
    event.listen(process_engine,"before_cursor_execute",record)
    try:response=client.post("/api/v1/exports/kitchen-operations/xlsx",headers=headers,json={"menu_id":menu["id"]})
    finally:event.remove(process_engine,"before_cursor_execute",record)
    assert response.status_code==200
    assert len([sql for sql in statements if sql.lstrip().upper().startswith("SELECT")])<=4

def test_menu_full_and_pretty_exports_are_authenticated_binary_downloads(client,db_session):
    headers=auth(client,db_session);category,dishes=foundations(client,headers);value=menu(client,headers,category["id"],name="中文匯出菜單",start="2026-09-01",end="2026-09-07");meal_types=meals(client,headers,value["id"],("早餐","早點","午餐","午點","晚餐","晚點","特殊餐"))
    assert client.put(f"/api/v1/menus/{value['id']}/editor",headers=headers,json=full_structure(value["id"],meal_types,dishes,days=7)).status_code==200
    for layout in ("full","grid","pretty"):
        xlsx=client.get(f"/api/v1/exports/menus/{value['id']}/{layout}/xlsx",headers=headers)
        assert xlsx.status_code==200,xlsx.text;assert xlsx.content[:2]==b"PK";assert "attachment" in xlsx.headers["content-disposition"]
        assert load_workbook(BytesIO(xlsx.content)).sheetnames
        pdf=client.get(f"/api/v1/exports/menus/{value['id']}/{layout}/pdf",headers=headers)
        assert pdf.status_code==200,pdf.text;reader=PdfReader(BytesIO(pdf.content));assert reader.pages
        assert "".join(page.extract_text() or "" for page in reader.pages)==""
    for layout in ("full","grid"):
        poster=client.get(f"/api/v1/exports/menus/{value['id']}/{layout}/xlsx?variant=poster",headers=headers)
        assert poster.status_code==200,poster.text
        sheet=load_workbook(BytesIO(poster.content)).active
        assert len(sheet.col_breaks.brk)==1 and len(sheet.row_breaks.brk)==1 and sheet.page_setup.pageOrder=="overThenDown"
    assert client.get(f"/api/v1/exports/menus/{value['id']}/pretty/xlsx?variant=poster",headers=headers).status_code==422

def test_menu_exports_require_auth(client):
    response=client.get("/api/v1/exports/menus/00000000-0000-0000-0000-000000000001/full/xlsx")
    assert response.status_code==401

def test_menu_export_query_budget_is_constant(client,db_session):
    headers=auth(client,db_session);category,dishes=foundations(client,headers);value=menu(client,headers,category["id"]);meal_types=meals(client,headers,value["id"])
    client.put(f"/api/v1/menus/{value['id']}/editor",headers=headers,json=full_structure(value["id"],meal_types,dishes));statements=[]
    def record(conn,cursor,statement,parameters,context,executemany):statements.append(statement)
    event.listen(process_engine,"before_cursor_execute",record)
    try:response=client.get(f"/api/v1/exports/menus/{value['id']}/full/xlsx",headers=headers)
    finally:event.remove(process_engine,"before_cursor_execute",record)
    assert response.status_code==200
    assert len([sql for sql in statements if sql.lstrip().upper().startswith("SELECT")])<=4


def test_menu_nutrition_exports_are_opt_in_complete_and_deduplicated(client,db_session):
    headers=auth(client,db_session);category,dishes=foundations(client,headers)
    ingredient_category=client.post("/api/v1/categories/ingredient",headers=headers,json={"name":"匯出營養食材"}).json()
    food=client.post("/api/v1/nutrition/manual-foods",headers=headers,json={"name":"匯出雞肉營養","nutrients":{"corrected_energy":"200","protein":"20","fat":"2","sodium":"50"}}).json()
    ingredient=client.post("/api/v1/ingredients",headers=headers,json={"code":"EXPORT-N","name":"匯出營養雞肉","category_id":ingredient_category["id"],"unit":"g","current_price":"1"}).json()
    client.patch(f"/api/v1/ingredients/{ingredient['id']}/nutrition",headers=headers,json={"nutrition_food_id":food["id"]})
    client.put(f"/api/v1/dishes/{dishes[0]['id']}/recipe",headers=headers,json={"items":[{"ingredient_id":ingredient["id"],"quantity":"150","unit":"g","loss_rate":"0"}]})
    value=menu(client,headers,category["id"],name="營養匯出菜單");meal_types=meals(client,headers,value["id"])
    client.put(f"/api/v1/menus/{value['id']}/editor",headers=headers,json=full_structure(value["id"],meal_types,dishes))

    original=client.get(f"/api/v1/exports/menus/{value['id']}/full/xlsx",headers=headers)
    original_text="\n".join(str(cell.value) for sheet in load_workbook(BytesIO(original.content)) for row in sheet.iter_rows() for cell in row)
    assert "kcal" not in original_text

    statements=[]
    def record(conn,cursor,statement,parameters,context,executemany):statements.append(statement)
    event.listen(process_engine,"before_cursor_execute",record)
    try:calories=client.get(f"/api/v1/exports/menus/{value['id']}/full/xlsx?nutrition=calories",headers=headers)
    finally:event.remove(process_engine,"before_cursor_execute",record)
    assert calories.status_code==200,calories.text
    calorie_text="\n".join(str(cell.value) for sheet in load_workbook(BytesIO(calories.content)) for row in sheet.iter_rows() for cell in row)
    assert "測試菜色1（300 kcal）" in calorie_text and "測試菜色2（無）" in calorie_text
    assert len([sql for sql in statements if sql.lstrip().upper().startswith("SELECT")])<=8

    detailed=client.get(f"/api/v1/exports/menus/{value['id']}/grid/xlsx?nutrition=detailed",headers=headers)
    workbook=load_workbook(BytesIO(detailed.content));assert workbook.sheetnames==["原版-菜色分格週表","熱量-菜色分格週表","詳細營養"]
    assert workbook["詳細營養"].max_row==4 and workbook["詳細營養"].max_column==10  # dish + all nine reportable definitions
    assert workbook["詳細營養"].freeze_panes=="B3" and workbook["詳細營養"].auto_filter.ref=="A2:J4"
    assert workbook["詳細營養"].cell(2,2).value=="修正熱量 (kcal)" and workbook["詳細營養"].cell(2,3).value=="熱量 (kcal)"
    detail_values={workbook["詳細營養"].cell(row,1).value:workbook["詳細營養"].cell(row,2).value for row in range(3,5)}
    assert detail_values=={"測試菜色1":300,"測試菜色2":"無"}
    pdf=client.get(f"/api/v1/exports/menus/{value['id']}/pretty/pdf?nutrition=detailed",headers=headers)
    assert pdf.status_code==200,pdf.text;reader=PdfReader(BytesIO(pdf.content));assert len(reader.pages)>=3
    assert all((page.extract_text() or "")=="" for page in reader.pages)
    assert client.get(f"/api/v1/exports/menus/{value['id']}/full/xlsx?variant=poster&nutrition=calories",headers=headers).status_code==200
    assert client.get(f"/api/v1/exports/menus/{value['id']}/full/xlsx?variant=poster&nutrition=detailed",headers=headers).status_code==422

def test_kitchen_simple_excel_pdf_known_answer_anomaly_and_query_budget(client,db_session):
    headers=auth(client,db_session);value,*_=kitchen_fixture(client,headers,db_session);body={"menu_id":value["id"],"selected_dates":["2026-10-01"]};statements=[]
    def record(conn,cursor,statement,parameters,context,executemany):statements.append(statement)
    event.listen(process_engine,"before_cursor_execute",record)
    try:xlsx=client.post("/api/v1/exports/kitchen-operations/simple/xlsx",headers=headers,json=body)
    finally:event.remove(process_engine,"before_cursor_execute",record)
    assert xlsx.status_code==200,xlsx.text
    workbook=load_workbook(BytesIO(xlsx.content));values=[cell.value for sheet in workbook for row in sheet.iter_rows() for cell in row]
    text="\n".join(str(value) for value in values)
    assert "雞肉" in text and "16.5 kg" in text and "⚠ 此菜有配方資料異常，請確認" in text
    assert len([sql for sql in statements if sql.lstrip().upper().startswith("SELECT")])<=4
    pdf=client.post("/api/v1/exports/kitchen-operations/simple/pdf",headers=headers,json=body)
    assert pdf.status_code==200,pdf.text
    assert "".join(page.extract_text() or "" for page in PdfReader(BytesIO(pdf.content)).pages)==""
    poster=client.post("/api/v1/exports/kitchen-operations/simple/xlsx?variant=poster",headers=headers,json=body)
    assert poster.status_code==200 and len(load_workbook(BytesIO(poster.content)).active.row_breaks.brk)==1
    assert client.post("/api/v1/exports/kitchen-operations/simple/pdf?variant=poster",headers=headers,json=body).status_code==422

def test_kitchen_simple_excludes_inactive_dynamic_meal_type(client,db_session):
    headers=auth(client,db_session);value,lunch,dinner,*_=kitchen_fixture(client,headers,db_session)
    assert client.post(f"/api/v1/menus/{value['id']}/meal-types/{dinner['id']}/deactivate",headers=headers).status_code==200
    response=client.post("/api/v1/exports/kitchen-operations/simple/xlsx",headers=headers,json={"menu_id":value["id"],"selected_dates":["2026-10-01"]})
    assert response.status_code==200,response.text
    values=[cell.value for sheet in load_workbook(BytesIO(response.content)) for row in sheet.iter_rows() for cell in row]
    assert any(isinstance(item,str) and "午餐" in item for item in values)
    assert not any(isinstance(item,str) and "晚餐" in item for item in values) and "高湯" not in values
