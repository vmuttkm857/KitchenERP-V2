from datetime import date
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from app.domains.exports.excel import kitchen_workbook
from app.domains.exports.kitchen_a4 import (
    kitchen_a4_workbook,
    paginate_dishes,
    safe_worksheet_title,
)


def ingredient(name="紅蘿蔔", quantity="12.5", per_person="0.125"):
    return {
        "ingredient_name": name,
        "display_quantity": Decimal(quantity),
        "display_unit": "kg",
        "quantity_per_person": Decimal(per_person),
        "recipe_unit": "kg",
    }


def result():
    dishes = [
        {"dish_code": "01", "dish_name": "燉菜", "ingredients": [ingredient(), ingredient("馬鈴薯", "8", "0.08")]},
        {"dish_code": "02", "dish_name": "清湯", "ingredients": [ingredient("青菜", "3", "0.03")]},
    ]
    return {
        "menu": {"menu_name": "中文菜單"},
        "days": [
            {
                "menu_date": date(2026, 8, 28),
                "meals": [
                    {"meal_type_name": "午餐", "dishes": dishes},
                    {"meal_type_name": "晚餐", "dishes": dishes[:1]},
                ],
            },
            {
                "menu_date": date(2026, 8, 29),
                "meals": [{"meal_type_name": "午餐", "dishes": dishes[:1]}],
            },
        ],
        "ingredient_summary": [],
        "anomalies": [],
    }


def test_a4_groups_each_date_and_meal_into_an_independent_print_sheet():
    wb = load_workbook(BytesIO(kitchen_a4_workbook(result())))
    assert len(wb.sheetnames) == 3
    assert wb.sheetnames == ["08_28_午餐", "08_28_晚餐", "08_29_午餐"]
    ws = wb["08_28_午餐"]
    assert ws["A1"].value == "廚房作業備料表"
    assert ws["A2"].value == "菜單：中文菜單"
    assert ws["C2"].value == "日期：2026/08/28"
    assert ws["A3"].value == "午餐"
    assert ws["C3"].value == "第 1 / 1 頁"
    assert ws.oddHeader.center.text is None


def test_a4_column_order_numeric_cells_and_print_settings():
    ws = load_workbook(BytesIO(kitchen_a4_workbook(result())))["08_28_午餐"]
    assert [ws.cell(6, column).value for column in range(1, 7)] == ["食材名稱", "備料量", "單位", None, "每人量", "單位"]
    assert ws["A7"].value == "紅蘿蔔"
    assert isinstance(ws["B7"].value, (int, float)) and ws["B7"].value == 12.5
    assert isinstance(ws["E7"].value, (int, float)) and ws["E7"].value == 0.125
    assert ws["B7"].number_format == "#,##0.######"
    assert ws["B8"].number_format == "#,##0"
    assert ws["B7"].alignment.horizontal == "right" and ws["B7"].alignment.vertical == "center"
    assert ws["C7"].alignment.horizontal == "left" and ws["C7"].alignment.vertical == "center"
    assert ws["E7"].alignment.horizontal == "right" and ws["E7"].alignment.vertical == "center"
    assert ws["F7"].alignment.horizontal == "left" and ws["F7"].alignment.vertical == "center"
    widths = [ws.column_dimensions[column].width for column in "ABCDEF"]
    assert widths == [37, 17, 8, 6.5, 17, 8]
    assert round(sum(widths), 2) == 93.5
    assert sum(widths) / 75.4 >= 1.20
    assert ws["C6"].value == "單位" and ws["C6"].alignment.wrap_text is not True
    assert ws["F6"].value == "單位" and ws["F6"].alignment.wrap_text is not True
    assert ws.column_dimensions["C"].width >= 7
    assert ws.column_dimensions["F"].width >= 7
    assert ws.page_setup.paperSize == 9
    assert ws.page_setup.orientation == "portrait"
    assert ws.page_setup.fitToWidth == 1 and ws.page_setup.fitToHeight == 0
    assert ws.page_setup.scale is None
    assert ws.page_margins.left == 0.25 and ws.page_margins.right == 0.25
    assert ws.page_margins.top == 0.35 and ws.page_margins.bottom == 0.35
    assert len(ws.col_breaks.brk) == 0
    assert ws.print_options.horizontalCentered is True
    assert str(ws.print_area).endswith(f"$A$1:$F${ws.max_row}")
    assert not any(cell.value == 10 for row in ws.iter_rows() for cell in row)


def test_a4_page_header_uses_final_readable_font_sizes():
    ws = load_workbook(BytesIO(kitchen_a4_workbook(result())))["08_28_午餐"]
    assert ws["A1"].font.sz == 20 and ws["A1"].font.bold
    assert ws["A2"].font.sz == 17 and ws["A2"].font.bold
    assert ws["C2"].font.sz == 17 and ws["C2"].font.bold
    assert ws["A3"].font.sz == 20 and ws["A3"].font.bold
    assert ws["C3"].font.sz == 16 and ws["C3"].font.bold


def test_a4_header_reserves_non_overlapping_space_for_full_date_and_page_number():
    ws = load_workbook(BytesIO(kitchen_a4_workbook(result())))["08_28_午餐"]
    merged = {str(cell_range) for cell_range in ws.merged_cells.ranges}
    assert "A2:B2" in merged and "C2:F2" in merged
    assert "A3:B3" in merged and "C3:F3" in merged
    assert ws["C2"].value == "日期：2026/08/28"
    assert ws["C2"].alignment.wrap_text is not True
    assert ws["C3"].value == "第 1 / 1 頁"
    assert ws["C3"].alignment.wrap_text is not True


def test_pagination_keeps_normal_dish_together_and_marks_oversized_continuation():
    normal = {"dish_name": "一般菜", "ingredients": [ingredient(str(index)) for index in range(5)]}
    oversized = {"dish_name": "大型菜", "ingredients": [ingredient(str(index)) for index in range(35)]}
    pages = paginate_dishes([normal, normal], rows_per_page=10)
    assert len(pages) == 2 and [len(page) for page in pages] == [1, 1]
    large_pages = paginate_dishes([oversized], rows_per_page=10)
    assert len(large_pages) > 1
    assert large_pages[0][0]["continuation"] is False
    assert all(page[0]["continuation"] is True for page in large_pages[1:])
    assert sum(len(page[0]["ingredients"]) for page in large_pages) == 35


def test_sheet_names_are_safe_unique_and_deterministic():
    used = set()
    first = safe_worksheet_title("超長/餐別:*?[]\\名稱" * 3, used)
    second = safe_worksheet_title("超長/餐別:*?[]\\名稱" * 3, used)
    assert len(first) <= 31 and len(second) <= 31
    assert not any(character in first for character in "[]:*?/\\")
    assert second.endswith("_2") and first != second


def test_oversized_dish_creates_manual_page_breaks_with_repeated_dish_identity():
    data = result()
    data["days"] = [{"menu_date": date(2026, 8, 28), "meals": [{"meal_type_name": "午餐", "dishes": [{"dish_code": "99", "dish_name": "大型菜", "ingredients": [ingredient(str(index)) for index in range(40)]}]}]}]
    ws = load_workbook(BytesIO(kitchen_a4_workbook(data)))["08_28_午餐"]
    assert len(ws.row_breaks.brk) == 1
    second_page_row = ws.row_breaks.brk[0].id + 1
    assert ws["C3"].value == "第 1 / 2 頁"
    assert ws.cell(second_page_row + 2, 3).value == "第 2 / 2 頁"
    assert ws.cell(second_page_row, 1).value == "廚房作業備料表"
    assert ws.cell(second_page_row + 1, 1).value == "菜單：中文菜單"
    assert ws.cell(second_page_row + 1, 3).value == "日期：2026/08/28"
    assert ws.cell(second_page_row + 2, 1).value == "午餐"
    dish_titles = [cell.value for cell in ws["A"] if isinstance(cell.value, str) and "大型菜" in cell.value]
    assert dish_titles[0] == "【99 大型菜】"
    assert all("（續）" in title for title in dish_titles[1:])


def test_formula_like_names_are_escaped_and_original_export_is_unchanged():
    data = result()
    data["days"][0]["meals"][0]["dishes"][0]["ingredients"][0]["ingredient_name"] = "=cmd"
    ws = load_workbook(BytesIO(kitchen_a4_workbook(data)))["08_28_午餐"]
    assert ws["A7"].value == "'=cmd"
    original = load_workbook(BytesIO(kitchen_workbook({
        "menu": {"menu_name": "中文菜單"},
        "days": [{"menu_date": date(2026, 8, 28), "meals": [{"meal_type_name": "午餐", "dishes": [{"dish_code": "D1", "dish_name": "燉菜", "diner_count": 10, "ingredients": [{"ingredient_code": "I1", "ingredient_name": "紅蘿蔔", "display_quantity": Decimal("1"), "display_unit": "kg", "notes": None, "anomalies": []}]}]}]}],
        "ingredient_summary": [],
        "anomalies": [],
    })))
    assert original.sheetnames == ["備料明細", "食材彙總", "異常"]
