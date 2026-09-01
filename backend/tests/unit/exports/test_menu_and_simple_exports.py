from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook
from pypdf import PdfReader

from app.domains.exports.kitchen_simple import WARNING_TEXT, display_ingredient, kitchen_simple_pdf, kitchen_simple_workbook, simple_page_plan
from app.domains.exports.menu_exports import menu_full_pdf, menu_full_workbook, menu_grid_pdf, menu_grid_workbook, menu_nutrition_pdf, menu_nutrition_workbook, menu_pretty_pdf, menu_pretty_workbook, menu_week_plan
from app.domains.nutrition.calculator import NutrientDefinition


def menu_result(days=7, meal_count=3, dishes_per_slot=2):
    dates=[date(2026,8,31)+timedelta(days=i) for i in range(days)]
    meals=[SimpleNamespace(id=f"meal-{i}",name=f"自訂餐別{i+1}",sort_order=i+1,is_active=True) for i in range(meal_count)]
    slots=[{"menu_date":day,"menu_meal_type_id":meal.id,"dishes":[{"dish_id":f"dish-{j}","dish_name":"=長菜名"+"香菇雞腿"*(j+1),"diner_count":100+j,"sort_order":j+1} for j in range(dishes_per_slot)]} for day in dates for meal in meals]
    return {"menu":{"name":"中文測試菜單"},"dates":dates,"meal_types":meals,"slots":slots}


def kitchen_result(meal_count=3):
    error={"severity":"error","code":"ZERO_RECIPE_QUANTITY","message":"technical"}
    dishes=[{"dish_name":"香菇雞","diner_count":100,"recipe_ready":False,"anomalies":[],"ingredients":[
        {"ingredient_name":"雞腿","required_quantity":"12000","required_unit":"g","anomalies":[]},
        {"ingredient_name":"米酒","required_quantity":"1500","required_unit":"ml","anomalies":[]},
        {"ingredient_name":"薑","required_quantity":None,"required_unit":"g","anomalies":[error]},
    ]}]
    meals=[{"meal_type_id":f"meal-{i}","meal_type_name":f"餐別{i+1}","dishes":dishes} for i in range(meal_count)]
    return {"menu":{"menu_name":"廚房中文菜單"},"days":[{"menu_date":date(2026,8,31)+timedelta(days=i),"meals":meals} for i in range(7)]}


def values(sheet): return [cell.value for row in sheet.iter_rows() for cell in row]


def assert_single_a4(sheet):
    assert str(sheet.page_setup.paperSize)=="9"; assert sheet.page_setup.orientation=="landscape"
    assert sheet.page_setup.fitToWidth==1 and sheet.page_setup.fitToHeight==1; assert sheet.print_area=="'"+sheet.title+"'!$A$1:$H$"+str(sheet.max_row)


def assert_image_only(payload):
    reader=PdfReader(BytesIO(payload)); assert len(reader.pages)==1
    page=reader.pages[0]; assert float(page.mediabox.width)>float(page.mediabox.height); assert (page.extract_text() or "")==""
    assert page["/Resources"].get("/XObject"); content=page.get_contents().get_data(); assert b"Tj" not in content and b"TJ" not in content


@pytest.mark.parametrize("meal_count",[3,5,6])
def test_merged_week_is_dynamic_single_a4_without_diner_counts(meal_count):
    workbook=load_workbook(BytesIO(menu_full_workbook(menu_result(meal_count=meal_count)))); sheet=workbook.active; assert_single_a4(sheet)
    assert sheet.max_column==8 and sheet.max_row==meal_count+2; assert "\n" in sheet["B3"].value; assert "100人" not in "".join(str(v) for v in values(sheet))


def test_grid_week_keeps_one_dish_per_cell_and_unmerged_meal_rows():
    sheet=load_workbook(BytesIO(menu_grid_workbook(menu_result(dishes_per_slot=4)))).active; assert_single_a4(sheet)
    assert sheet.max_row==14; assert sheet["A3"].value=="自訂餐別1" and sheet["A4"].value is None
    assert sheet["B3"].value.startswith("'=長菜名") and "\n" not in sheet["B3"].value
    assert all(str(region)=="A1:H1" for region in sheet.merged_cells.ranges)


def test_pretty_week_remains_one_landscape_grid_without_metadata():
    sheet=load_workbook(BytesIO(menu_pretty_workbook(menu_result(meal_count=6,dishes_per_slot=6)))).active; assert_single_a4(sheet)
    text="".join(str(v) for v in values(sheet)); assert "100人" not in text and "第 1 頁" not in text and sheet.max_column==8


@pytest.mark.parametrize("builder",[menu_full_pdf,menu_grid_pdf,menu_pretty_pdf])
def test_menu_pdfs_are_single_landscape_image_pages(builder): assert_image_only(builder(menu_result(meal_count=6,dishes_per_slot=6)))


@pytest.mark.parametrize("builder",[menu_full_workbook,menu_grid_workbook])
def test_menu_poster_has_manual_two_by_two_breaks(builder):
    sheet=load_workbook(BytesIO(builder(menu_result(meal_count=6,dishes_per_slot=4),"poster"))).active
    assert sheet.page_setup.fitToWidth==0 and sheet.page_setup.fitToHeight==0 and sheet.page_setup.pageOrder=="overThenDown"
    assert [b.id for b in sheet.col_breaks.brk]==[4] and len(sheet.row_breaks.brk)==1; assert sheet.print_area.endswith(f"$H${sheet.max_row}")


def test_kitchen_display_conversion_is_report_only_and_safe():
    assert display_ingredient({"required_quantity":"1000","required_unit":"g"})==("1","kg")
    assert display_ingredient({"required_quantity":"2","required_unit":"斤"})==("1.2","kg")
    assert display_ingredient({"required_quantity":"1500","required_unit":"ml"})==("1.5","L")
    assert display_ingredient({"required_quantity":"3","required_unit":"盒"})==("3","盒")


def test_kitchen_week_keeps_hierarchy_safe_rows_and_human_warning():
    plan=simple_page_plan(kitchen_result()); assert len(plan)==1 and len(plan[0]["dates"])==7
    sheet=load_workbook(BytesIO(kitchen_simple_workbook(kitchen_result()))).active; assert_single_a4(sheet)
    text="\n".join(str(v) for v in values(sheet)); assert "香菇雞　100人" in text and "雞腿　12 kg" in text and "米酒　1.5 L" in text
    assert WARNING_TEXT in text and "ZERO_RECIPE_QUANTITY" not in text and "薑　不可計算 g" in text


def test_kitchen_poster_and_pdf_print_contracts():
    poster=load_workbook(BytesIO(kitchen_simple_workbook(kitchen_result(6),"poster"))).active
    assert [b.id for b in poster.col_breaks.brk]==[4] and len(poster.row_breaks.brk)==1 and poster.page_setup.pageOrder=="overThenDown"
    assert_image_only(kitchen_simple_pdf(kitchen_result(6)))


def test_weeks_are_never_split_inside_seven_day_period():
    plans=menu_week_plan(menu_result(days=11,meal_count=6,dishes_per_slot=6)); assert [len(p["dates"]) for p in plans]==[7,4]


def nutrient_definitions(count=20):
    common=[("corrected_energy","修正熱量","kcal"),("energy","熱量","kcal"),("protein","粗蛋白","g"),("fat","粗脂肪","g"),("carbohydrate","總碳水化合物","g"),("dietary_fiber","膳食纖維","g"),("sodium","鈉","mg"),("potassium","鉀","mg"),("calcium","鈣","mg")]
    definitions=[NutrientDefinition(*item) for item in common]
    definitions.extend(NutrientDefinition(f"synthetic_{index:03}",f"營養素 {index:03}","mg") for index in range(max(0,count-len(definitions))))
    return tuple(definitions)


def nutrition_results(result, definitions=None):
    definitions=definitions or nutrient_definitions()
    values={"corrected_energy":"350","protein":"28.4","fat":"18.1","carbohydrate":None,"dietary_fiber":"0","sodium":"560","potassium":"300","calcium":"25"}
    output={}
    for dish_id,_ in {(item["dish_id"],item["dish_name"]) for slot in result["slots"] for item in slot["dishes"]}:
        nutrients={definition.code:SimpleNamespace(code=definition.code,name=definition.name,unit=definition.unit,value=Decimal("0") if definition.code.startswith("synthetic_") else (None if values.get(definition.code) is None else Decimal(values[definition.code])),complete=definition.code.startswith("synthetic_") or values.get(definition.code) is not None) for definition in definitions}
        output[dish_id]=SimpleNamespace(calorie_complete=True,calorie_value=Decimal("350"),nutrients=nutrients)
    return output


@pytest.mark.parametrize("layout",["merged","grid","pretty"])
def test_calorie_workbook_preserves_layout_and_adds_kcal(layout):
    result=menu_result(); definitions=nutrient_definitions(); workbook=load_workbook(BytesIO(menu_nutrition_workbook(result,layout,"calories","single",nutrition_results(result,definitions),definitions)))
    text="\n".join(str(value) for sheet in workbook for value in values(sheet))
    assert "350 kcal" in text and len(workbook.sheetnames)==1


def test_detailed_workbook_has_original_calorie_and_deduplicated_native_detail():
    result=menu_result(); definitions=nutrient_definitions(105); workbook=load_workbook(BytesIO(menu_nutrition_workbook(result,"merged","detailed","single",nutrition_results(result,definitions),definitions)))
    assert workbook.sheetnames==["原版-餐別合併週表","熱量-餐別合併週表","詳細營養"]
    assert "kcal" not in "\n".join(str(value) for value in values(workbook.worksheets[0]))
    assert "350 kcal" in "\n".join(str(value) for value in values(workbook.worksheets[1]))
    detail=workbook["詳細營養"]; assert detail.max_row==4 and detail.max_column==106
    assert detail["B3"].value==350 and detail["C3"].value=="無" and detail["F3"].value=="無" and detail["G3"].value==0
    assert detail["B3"].data_type=="n" and detail["K3"].value==0 and detail["K3"].data_type=="n"
    assert detail.freeze_panes=="B3" and detail.print_title_cols=="$A:$A" and detail.auto_filter.ref=="A2:DB4"
    assert [detail.cell(2,column).value for column in range(2,7)]==["修正熱量 (kcal)","熱量 (kcal)","粗蛋白 (g)","粗脂肪 (g)","總碳水化合物 (g)"]
    assert detail.page_setup.fitToWidth==0


def test_nutrition_pdf_remains_image_only_and_detail_can_extend_pages():
    result=menu_result(dishes_per_slot=12); definitions=nutrient_definitions(105); payload=menu_nutrition_pdf(result,"pretty","detailed",nutrition_results(result,definitions),definitions)
    reader=PdfReader(BytesIO(payload)); assert len(reader.pages)==32  # original + calorie + 15 nutrient chunks x 2 dish pages
    assert all((page.extract_text() or "")=="" for page in reader.pages)
