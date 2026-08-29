from datetime import date,datetime
from decimal import Decimal
from io import BytesIO
from openpyxl import load_workbook
from pypdf import PdfReader
from app.domains.exports.excel import kitchen_workbook,requirements_workbook
from app.domains.exports.excel import purchase_workbook,snapshot_workbook
from app.domains.exports.pdf import kitchen_pdf,paginate_blocks,purchase_pdf
from app.domains.exports.safety import content_disposition,safe_cell_text,safe_filename

def requirement_result(name="繁體食材"):
    row={"row_key":"1","ingredient_code":"I01","ingredient_name":name,"supplier_name":"台灣供應商","requirement_quantity":Decimal("0.12345678"),"requirement_unit":"kg","suggested_purchase_quantity":Decimal("0.2"),"suggested_purchase_unit":"kg","current_price":Decimal("12.34"),"estimated_cost":Decimal("2.468"),"needs_review":False}
    daily=[
        {"requirement_date":date(2026,10,6),"menu_id":"m2","menu_name":"菜單乙","supplier_name":"台灣供應商","ingredient_id":"i1","ingredient_code":"I01","ingredient_name":name,"quantity":Decimal("2.5"),"unit":"kg"},
        {"requirement_date":date(2026,10,5),"menu_id":"m1","menu_name":"菜單甲","supplier_name":"台灣供應商","ingredient_id":"i1","ingredient_code":"I01","ingredient_name":name,"quantity":Decimal("1.25"),"unit":"kg"},
    ]
    return {"rows":[row],"daily_rows":daily,"supplier_groups":[{"supplier_name":"台灣供應商","row_keys":["1"]}],"anomalies":[]}
def kitchen_result():
    anomaly={"severity":"warning","code":"CHECK","message":"請確認異常","related_entity_name":"食材"}
    line={"ingredient_code":"I01","ingredient_name":"紅蘿蔔","display_quantity":Decimal("1.25"),"display_unit":"kg","notes":"備料","anomalies":[anomaly]}
    dish={"dish_code":"D01","dish_name":"燉菜","diner_count":10,"ingredients":[line]}
    return {"menu":{"menu_name":"中文菜單"},"days":[{"menu_date":date(2026,8,28),"meals":[{"meal_type_name":"午餐","dishes":[dish]}]}],"ingredient_summary":[{"ingredient_code":"I01","ingredient_name":"紅蘿蔔","supplier_name":"供應商甲","display_quantity":Decimal("1.25"),"display_unit":"kg","source_count":1,"anomalies":[]}],"anomalies":[anomaly]}
def test_excel_keeps_chinese_numeric_precision_and_escapes_formulas():
    wb=load_workbook(BytesIO(requirements_workbook(requirement_result("=HYPERLINK(\"bad\")"))));ws=wb["需求彙總"]
    assert ws["B2"].value.startswith("'=");assert ws["C2"].value=="台灣供應商"
    assert Decimal(str(ws["D2"].value))==Decimal("0.12345678");assert isinstance(ws["D2"].value,(int,float))
    daily=wb["每日採購需求"]
    assert [cell.value for cell in daily[1]]==["使用日期","菜單","供應商","食材編號","食材名稱","需求量","單位"]
    assert daily["A2"].value.date()==date(2026,10,5) and daily["B2"].value=="菜單甲"
    assert isinstance(daily["F2"].value,(int,float)) and Decimal(str(daily["F2"].value))==Decimal("1.25")
    assert daily["E2"].value.startswith("'=")

def test_daily_requirement_excel_preserves_supplier_and_menu_attribution():
    result=requirement_result();result["daily_rows"].append({"requirement_date":date(2026,10,5),"menu_id":"m3","menu_name":"同日另一菜單","supplier_name":"另一供應商","ingredient_id":"i2","ingredient_code":"I02","ingredient_name":"青江菜","quantity":Decimal("3"),"unit":"kg"})
    ws=load_workbook(BytesIO(requirements_workbook(result)))["每日採購需求"]
    values=[tuple(cell.value for cell in row) for row in ws.iter_rows(min_row=2)]
    assert (datetime(2026,10,5),"同日另一菜單","另一供應商","I02","青江菜",3,"kg") in values
def test_kitchen_excel_has_all_sections():
    wb=load_workbook(BytesIO(kitchen_workbook(kitchen_result())))
    assert wb.sheetnames==["備料明細","食材彙總","異常"];assert wb["備料明細"]["D2"].value=="燉菜"
def test_pdf_embeds_chinese_font_and_is_readable():
    reader=PdfReader(BytesIO(kitchen_pdf(kitchen_result())))
    assert "廚房備料表" in "".join(page.extract_text() or "" for page in reader.pages)
    fonts=[]
    for page in reader.pages:fonts.extend(page["/Resources"]["/Font"].values())
    assert any("NotoSansTC" in str(font.get_object().get("/BaseFont")) for font in fonts)
def test_filename_and_formula_safety():
    assert safe_filename('採購:甲/乙*?',"xlsx")=="採購_甲_乙__.xlsx";assert safe_cell_text("+cmd")=="'+cmd";assert "filename*=UTF-8''" in content_disposition("中文.xlsx")
def test_pagination_is_pure_and_stable():
    rows=list(range(45));pages=paginate_blocks(rows,20)
    assert [len(page) for page in pages]==[20,20,5];assert [item for page in pages for item in page]==rows;assert paginate_blocks([],20)==[[]]

def test_snapshot_excel_preserves_hard_copy_fields():
    result={"criteria":{"menu_ids":["m1"]},"revision":2,"created_at":"2026-08-28T00:00:00Z","created_by_name":"管理者","source_menus":[{"menu_name":"週菜單"}],"total_estimated_cost":Decimal("10.50"),"known_estimated_cost":Decimal("10.50"),"items":[{"ingredient_code_snapshot":"I1","ingredient_name_snapshot":"青菜","supplier_name_snapshot":"供應商","requirement_quantity":Decimal("1.1"),"requirement_unit":"kg","suggested_purchase_quantity":Decimal("2"),"adjusted_quantity":Decimal("3"),"purchase_unit_snapshot":"kg","adjusted_estimated_cost":Decimal("10.50"),"unit_price_snapshot":Decimal("3.50"),"estimated_cost_snapshot":Decimal("7"),"anomaly_snapshot":[]}],"anomaly_snapshot":[]}
    wb=load_workbook(BytesIO(snapshot_workbook(result)))
    assert wb.sheetnames==["快照摘要","固定品項","異常快照"];assert wb["固定品項"]["G2"].value==3

def test_purchase_excel_pdf_support_multiple_suppliers():
    class Item:
        ingredient_code_snapshot="I1";ingredient_name_snapshot="米";final_purchase_quantity=Decimal("2.5");purchase_unit_snapshot="kg";package_size_snapshot=Decimal("1");minimum_order_quantity_snapshot=Decimal("1");unit_price_snapshot=Decimal("10");purchase_cost_snapshot=Decimal("25");anomaly_snapshot=[]
    orders=[{"supplier_name_snapshot":"供應商甲","known_total_cost":Decimal("25"),"total_cost":Decimal("25"),"items":[Item()]},{"supplier_name_snapshot":"供應商乙","known_total_cost":Decimal("25"),"total_cost":Decimal("25"),"items":[Item()]}]
    result={"purchase_number":"PO-2026-001","status":"confirmed","created_at":"2026-08-28","created_by_name":"管理者","source_snapshot_revision":1,"notes":"中文採購","total_cost":Decimal("50"),"known_total_cost":Decimal("50"),"orders":orders}
    wb=load_workbook(BytesIO(purchase_workbook(result)));assert wb.sheetnames==["採購摘要","供應商甲","供應商乙"]
    assert len(PdfReader(BytesIO(purchase_pdf(result))).pages)>=1
