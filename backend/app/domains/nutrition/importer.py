from __future__ import annotations

import hashlib
import json
import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from io import BytesIO

from openpyxl import load_workbook

from app.domains.nutrition.exceptions import NutritionImportError

REQUIRED_HEADERS = {"整合編號", "食品分類", "樣品名稱"}
# P/M/S is a composite ratio (for example 5.73/1.00/2.83), not a scalar
# nutrient value and therefore cannot be represented by NUMERIC safely.
META_HEADERS = {"整合編號", "食品分類", "樣品名稱", "內容物描述", "俗名", "廢棄率(%)", "P/M/S"}
CANONICAL = {
    "修正熱量(kcal)": "corrected_energy", "熱量(kcal)": "energy", "粗蛋白(g)": "protein",
    "粗脂肪(g)": "fat", "總碳水化合物(g)": "carbohydrate", "碳水化合物(g)": "carbohydrate",
    "膳食纖維(g)": "dietary_fiber", "鈉(mg)": "sodium", "鉀(mg)": "potassium", "鈣(mg)": "calcium",
}


@dataclass(frozen=True)
class ParsedNutrient:
    code: str; name: str; unit: str | None; original_source_name: str; sort_order: int


@dataclass(frozen=True)
class ParsedFood:
    external_code: str; name: str; category: str | None; description: str | None
    aliases: list[str] | None; waste_rate: Decimal | None; values: dict[str, Decimal]; source_hash: str


@dataclass(frozen=True)
class ParsedWorkbook:
    header_row: int; nutrients: list[ParsedNutrient]; foods: list[ParsedFood]; errors: list[dict]; source_hash: str


def _text(value) -> str | None:
    if value is None: return None
    result = str(value).strip()
    return result or None


def _decimal(value, *, row: int, column: str, errors: list[dict]) -> Decimal | None:
    if value is None or (isinstance(value, str) and not value.strip()): return None
    if isinstance(value, bool):
        errors.append({"row": row, "column": column, "value": str(value)}); return None
    if isinstance(value, str) and value.startswith("="):
        errors.append({"row": row, "column": column, "value": "formula"}); return None
    try: return Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        errors.append({"row": row, "column": column, "value": str(value)[:80]}); return None


def _unit(header: str) -> str | None:
    match = re.search(r"\(([^()]*)\)\s*$", header)
    return match.group(1).strip() if match else None


def _code(header: str) -> str:
    if header in CANONICAL: return CANONICAL[header]
    return "tfda_" + hashlib.sha1(header.strip().encode("utf-8")).hexdigest()[:16]


def _hash(food: dict) -> str:
    return hashlib.sha256(json.dumps(food, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")).hexdigest()


def parse_tfda_xlsx(payload: bytes) -> ParsedWorkbook:
    if not payload or len(payload) > 20 * 1024 * 1024: raise NutritionImportError("XLSX 檔案為空或超過 20 MB")
    try: workbook = load_workbook(BytesIO(payload), read_only=True, data_only=False, keep_vba=False)
    except Exception as exc: raise NutritionImportError("無法讀取 XLSX 檔案") from exc
    if not workbook.sheetnames: raise NutritionImportError("XLSX 沒有工作表")
    sheet = workbook[workbook.sheetnames[0]]; header_row = None; headers = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 20), values_only=True), 1):
        candidate = [_text(value) for value in row]
        if REQUIRED_HEADERS.issubset({value for value in candidate if value}): header_row, headers = row_number, candidate; break
    if header_row is None: raise NutritionImportError("找不到必要欄位：整合編號、食品分類、樣品名稱")
    index = {header: position for position, header in enumerate(headers) if header}
    nutrient_headers = [header for header in headers if header and header not in META_HEADERS]
    nutrients = [ParsedNutrient(_code(header), re.sub(r"\([^()]*\)\s*$", "", header).strip(), _unit(header), header, order) for order, header in enumerate(nutrient_headers, 1)]
    errors: list[dict] = []; foods: list[ParsedFood] = []; seen: set[str] = set()
    for row_number, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
        external_code = _text(row[index["整合編號"]] if index["整合編號"] < len(row) else None)
        name = _text(row[index["樣品名稱"]] if index["樣品名稱"] < len(row) else None)
        if not external_code and not name: continue
        if not external_code or not name:
            errors.append({"row": row_number, "column": "整合編號/樣品名稱", "value": external_code or name or ""}); continue
        if external_code in seen:
            errors.append({"row": row_number, "column": "整合編號", "value": external_code}); continue
        seen.add(external_code)
        values = {}
        for nutrient in nutrients:
            position = index[nutrient.original_source_name]
            parsed = _decimal(row[position] if position < len(row) else None, row=row_number, column=nutrient.original_source_name, errors=errors)
            if parsed is not None: values[nutrient.code] = parsed
        waste = _decimal(row[index["廢棄率(%)"]] if "廢棄率(%)" in index and index["廢棄率(%)"] < len(row) else None, row=row_number, column="廢棄率(%)", errors=errors)
        aliases_text = _text(row[index["俗名"]]) if "俗名" in index and index["俗名"] < len(row) else None
        item = {"external_code": external_code, "name": name, "category": _text(row[index["食品分類"]]), "description": _text(row[index["內容物描述"]]) if "內容物描述" in index else None, "aliases": [part.strip() for part in re.split("[,，]", aliases_text) if part.strip()] if aliases_text else None, "waste_rate": str(waste) if waste is not None else None, "values": {key: str(value) for key, value in sorted(values.items())}}
        foods.append(ParsedFood(external_code, name, item["category"], item["description"], item["aliases"], waste, values, _hash(item)))
    if not foods: raise NutritionImportError("XLSX 沒有可匯入的食品資料")
    return ParsedWorkbook(header_row, nutrients, foods, errors, hashlib.sha256(payload).hexdigest())
