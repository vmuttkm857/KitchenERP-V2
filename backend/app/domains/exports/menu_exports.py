from __future__ import annotations

from datetime import date
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break
from PIL import Image, ImageDraw

from app.domains.exports.raster_pdf import GREEN, INK, LINE, PALE_GREEN, font, images_to_pdf, page_image, wrap_text
from app.domains.exports.safety import safe_cell_text
from app.domains.nutrition.calculator import NutrientDefinition

WEEKDAYS = "一二三四五六日"
THIN = Side(style="thin", color="C9D6CC")


def _date_text(value: date) -> str:
    return f"{value.month}/{value.day}（{WEEKDAYS[value.weekday()]}）"


def _display_nutrition(value: Decimal | None) -> str:
    if value is None: return "無"
    return format(value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP), "f").rstrip("0").rstrip(".")


def menu_week_plan(result: dict, nutrition_results: dict | None = None) -> list[dict]:
    used = {slot["menu_meal_type_id"] for slot in result["slots"]}
    labels_by_meal = {}
    for column in sorted(
        result.get("meal_type_columns", []),
        key=lambda item: (item.menu_meal_type_id, item.sort_order, str(item.id)),
    ):
        labels_by_meal.setdefault(column.menu_meal_type_id, []).append(str(column.name))
    meals = sorted(
        ({"id": meal.id, "name": str(meal.name), "sort_order": meal.sort_order,
          "labels": labels_by_meal.get(meal.id, [])}
         for meal in result["meal_types"] if meal.is_active or meal.id in used),
        key=lambda item: (item["sort_order"], str(item["id"])),
    )
    slots = {}
    for slot in result["slots"]:
        dishes = []
        for item in sorted(slot["dishes"], key=lambda item: (item["sort_order"], str(item["dish_id"]))):
            name = str(item["dish_name"])
            if nutrition_results is not None:
                nutrition = nutrition_results.get(item["dish_id"])
                calorie = _display_nutrition(nutrition.calorie_value) if nutrition and nutrition.calorie_complete else "無"
                name = f"{name}（{calorie}{' kcal' if calorie != '無' else ''}）"
            dishes.append(name)
        slots[(slot["menu_date"], slot["menu_meal_type_id"])] = dishes
    dates = list(result["dates"])
    return [{"dates": dates[start:start + 7], "meals": meals, "slots": slots} for start in range(0, len(dates), 7)] or [{"dates": [], "meals": meals, "slots": slots}]


def full_page_plan(result: dict) -> list[dict]: return menu_week_plan(result)
def pretty_page_plan(result: dict) -> list[dict]: return menu_week_plan(result)


def _density(plan: dict, grid: bool = False) -> tuple[float, float]:
    maximum = max((len(plan["slots"].get((day, meal["id"]), [])) for day in plan["dates"] for meal in plan["meals"]), default=1)
    pressure = len(plan["meals"]) + maximum * (1.25 if grid else .7)
    return (11, 42) if pressure <= 7 else ((9.5, 32) if pressure <= 11 else (8, 24))


def _sheet(workbook: Workbook, title: str):
    sheet = workbook.create_sheet(title[:31]); sheet.sheet_view.showGridLines = False
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4; sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    sheet.page_margins.left = sheet.page_margins.right = .18; sheet.page_margins.top = sheet.page_margins.bottom = .25
    sheet.column_dimensions["A"].width = 11
    sheet.column_dimensions["B"].width = 13
    for column in "CDEFGHI": sheet.column_dimensions[column].width = 19
    return sheet


def _single(sheet, last_row: int):
    sheet.page_setup.fitToWidth = sheet.page_setup.fitToHeight = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True; sheet.print_area = f"A1:I{last_row}"


def _poster(sheet, last_row: int, ends: list[int]):
    sheet.sheet_properties.pageSetUpPr.fitToPage = False; sheet.page_setup.scale = 175
    sheet.page_setup.fitToWidth = sheet.page_setup.fitToHeight = 0; sheet.page_setup.pageOrder = "overThenDown"
    sheet.print_area = f"A1:I{last_row}"; sheet.col_breaks.append(Break(id=5))
    sheet.row_breaks.append(Break(id=ends[(len(ends)-1)//2] if len(ends) > 1 else max(3, last_row//2)))


def _header(sheet, title: str, plan: dict, pretty: bool):
    sheet.merge_cells("A1:I1"); sheet["A1"] = safe_cell_text(title)
    sheet["A1"].font = Font(size=17 if pretty else 15, bold=True, color="244B32"); sheet["A1"].alignment = Alignment(horizontal="center", vertical="center"); sheet.row_dimensions[1].height = 30
    headers = ["餐別", "菜單欄位"] + [_date_text(day) for day in plan["dates"]] + [""] * (7-len(plan["dates"]))
    for column, value in enumerate(headers, 1):
        cell = sheet.cell(2, column, value); cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor="42725A" if pretty else "356B48")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    sheet.row_dimensions[2].height = 27


def _append_menu_sheets(workbook: Workbook, result: dict, layout: str, variant: str,
                        nutrition_results: dict | None = None, sheet_prefix: str = "") -> None:
    plans = menu_week_plan(result, nutrition_results)
    label = {"merged": "餐別合併週表", "grid": "菜色分格週表", "pretty": "漂亮公告版"}[layout]
    for index, plan in enumerate(plans, 1):
        title = f"{sheet_prefix}{label}" if len(plans) == 1 else f"{sheet_prefix}{label}-{index}"
        sheet = _sheet(workbook, title); suffix = " 含熱量" if nutrition_results is not None else ""
        _header(sheet, f'{result["menu"]["name"]} {label}{suffix}', plan, layout == "pretty")
        size, height = _density(plan, layout == "grid")
        if nutrition_results is not None: size, height = max(7, size - 1), height * 1.8
        if variant == "poster": size, height = size + 2, height * 1.45
        row, ends = 3, []
        for meal in plan["meals"]:
            counts = [len(plan["slots"].get((day, meal["id"]), [])) for day in plan["dates"]]
            rows = max(len(meal["labels"]), max(counts, default=0), 1)
            for offset in range(rows):
                sheet.cell(row, 1, safe_cell_text(meal["name"]) if offset == 0 else "")
                sheet.cell(row, 2, safe_cell_text(meal["labels"][offset]) if offset < len(meal["labels"]) else "")
                for day_index in range(7):
                    dishes = plan["slots"].get((plan["dates"][day_index], meal["id"]), []) if day_index < len(plan["dates"]) else []
                    sheet.cell(row, day_index + 3, safe_cell_text(dishes[offset]) if offset < len(dishes) else "")
                for column in range(1, 10):
                    cell = sheet.cell(row, column); cell.font = Font(size=size + (1 if column == 1 else 0), bold=column <= 2, color="244B32" if column <= 2 else INK.lstrip("#"))
                    cell.fill = PatternFill("solid", fgColor="F3F8F4" if column <= 2 else ("FAFCFA" if layout == "pretty" else "FFFFFF"))
                    cell.alignment = Alignment(horizontal="center" if column <= 2 or layout == "pretty" else "left", vertical="center", wrap_text=True); cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
                sheet.row_dimensions[row].height = max(18, height/rows); row += 1
            ends.append(row-1)
        last = max(2, row-1); _poster(sheet, last, ends) if variant == "poster" else _single(sheet, last)


def _ordered_dishes(result: dict) -> list[tuple[object, str]]:
    seen = set(); dishes = []
    for slot in result["slots"]:
        for item in sorted(slot["dishes"], key=lambda value: (value["sort_order"], str(value["dish_id"]))):
            if item["dish_id"] not in seen:
                seen.add(item["dish_id"]); dishes.append((item["dish_id"], str(item["dish_name"])))
    return dishes


def _append_nutrition_detail_sheet(
    workbook: Workbook, result: dict, nutrition_results: dict,
    nutrient_definitions: tuple[NutrientDefinition, ...],
) -> None:
    sheet = workbook.create_sheet("詳細營養"); sheet.sheet_view.showGridLines = False
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4; sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    sheet.page_setup.fitToWidth = 0; sheet.page_setup.fitToHeight = 0; sheet.sheet_properties.pageSetUpPr.fitToPage = False
    sheet.page_setup.pageOrder = "overThenDown"
    sheet.print_title_rows = "1:2"; sheet.print_title_cols = "A:A"; sheet.freeze_panes = "B3"
    last_column = get_column_letter(len(nutrient_definitions) + 1)
    sheet["A1"] = safe_cell_text(f'{result["menu"]["name"]} 詳細營養（每人配方）')
    sheet["A1"].font = Font(size=12, bold=True, color="244B32")
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 42
    headers = ["菜色"] + [f"{item.name} ({item.unit})" if item.unit else item.name for item in nutrient_definitions]
    for column, value in enumerate(headers, 1):
        cell = sheet.cell(2, column, value); cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor="356B48")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    sheet.column_dimensions["A"].width = 24
    for column in range(2, len(headers) + 1): sheet.column_dimensions[get_column_letter(column)].width = 15
    for row_index, (dish_id, name) in enumerate(_ordered_dishes(result), 3):
        sheet.cell(row_index, 1, safe_cell_text(name))
        nutrition = nutrition_results[dish_id]
        for column, definition in enumerate(nutrient_definitions, 2):
            nutrient = nutrition.nutrients[definition.code]
            cell = sheet.cell(row_index, column)
            if nutrient.complete:
                cell.value = nutrient.value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                cell.number_format = "0.##"
            else: cell.value = "無"
        for column in range(1, len(headers) + 1):
            cell = sheet.cell(row_index, column); cell.alignment = Alignment(horizontal="center" if column > 1 else "left", vertical="center", wrap_text=True)
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        sheet.row_dimensions[row_index].height = 27
    sheet.auto_filter.ref = f"A2:{last_column}{max(2, sheet.max_row)}"
    sheet.print_area = f"A1:{last_column}{max(2, sheet.max_row)}"


def menu_workbook(result: dict, layout: str = "merged", variant: str = "single") -> bytes:
    workbook = Workbook(); del workbook["Sheet"]; _append_menu_sheets(workbook, result, layout, variant)
    stream = BytesIO(); workbook.save(stream); return stream.getvalue()


def menu_nutrition_workbook(result: dict, layout: str, mode: str, variant: str, nutrition_results: dict,
                            nutrient_definitions: tuple[NutrientDefinition, ...]) -> bytes:
    workbook = Workbook(); del workbook["Sheet"]
    if mode == "detailed": _append_menu_sheets(workbook, result, layout, "single", sheet_prefix="原版-")
    _append_menu_sheets(workbook, result, layout, variant if mode == "calories" else "single", nutrition_results, "熱量-")
    if mode == "detailed": _append_nutrition_detail_sheet(workbook, result, nutrition_results, nutrient_definitions)
    stream = BytesIO(); workbook.save(stream); return stream.getvalue()


def menu_full_workbook(result: dict, variant: str = "single") -> bytes: return menu_workbook(result, "merged", variant)
def menu_grid_workbook(result: dict, variant: str = "single") -> bytes: return menu_workbook(result, "grid", variant)
def menu_pretty_workbook(result: dict) -> bytes: return menu_workbook(result, "pretty")


def _menu_images(result: dict, layout: str, nutrition_results: dict | None = None) -> list[Image.Image]:
    images: list[Image.Image] = []; label = {"merged": "餐別合併週表", "grid": "菜色分格週表", "pretty": "漂亮公告版"}[layout]
    for plan in menu_week_plan(result, nutrition_results):
        image = page_image("landscape"); draw = ImageDraw.Draw(image); margin, top = 65, 105
        suffix = " 含熱量" if nutrition_results is not None else ""
        draw.text((image.width//2, 42), f'{result["menu"]["name"]} {label}{suffix}', font=font(17), fill=GREEN, anchor="ma")
        meal_w, label_w = 115, 145; day_w = (image.width-margin*2-meal_w-label_w)/7; header_h = 60
        headers = ["餐別", "菜單欄位"] + [_date_text(day) for day in plan["dates"]] + [""]*(7-len(plan["dates"])); x = margin
        for i, value in enumerate(headers):
            width = meal_w if i == 0 else label_w if i == 1 else day_w; draw.rectangle((x, top, x+width, top+header_h), fill=GREEN, outline=LINE, width=2)
            draw.text((x+width/2, top+header_h/2), value, font=font(8.5), fill="white", anchor="mm"); x += width
        row_h = (image.height-top-header_h-55)/max(1, len(plan["meals"])); body_size = max(7, min(11, row_h/8)); y = top+header_h
        for meal in plan["meals"]:
            draw.rectangle((margin, y, margin+meal_w, y+row_h), fill=PALE_GREEN, outline=LINE, width=2); draw.text((margin+meal_w/2, y+row_h/2), meal["name"], font=font(body_size+1), fill=GREEN, anchor="mm")
            counts = [len(plan["slots"].get((day, meal["id"]), [])) for day in plan["dates"]]
            rows = max(len(meal["labels"]), max(counts, default=0), 1); part = row_h/rows
            for row_index in range(rows):
                row_top = y+row_index*part
                draw.rectangle((margin+meal_w, row_top, margin+meal_w+label_w, row_top+part), fill=PALE_GREEN, outline=LINE, width=2)
                label = meal["labels"][row_index] if row_index < len(meal["labels"]) else ""
                draw.text((margin+meal_w+label_w/2, row_top+part/2), label, font=font(body_size), fill=GREEN, anchor="mm")
            for day_i in range(7):
                left = margin+meal_w+label_w+day_i*day_w
                dishes = plan["slots"].get((plan["dates"][day_i], meal["id"]), []) if day_i < len(plan["dates"]) else []
                for row_index in range(rows):
                    row_top = y+row_index*part
                    draw.rectangle((left, row_top, left+day_w, row_top+part), fill="#FAFCFA" if layout == "pretty" else "white", outline=LINE, width=2)
                    if row_index < len(dishes):
                        lines = wrap_text(draw, dishes[row_index], font(body_size), day_w-18)
                        draw.multiline_text((left+day_w/2, row_top+part/2), "\n".join(lines), font=font(body_size), fill=INK, anchor="mm", align="center", spacing=3)
            y += row_h
        images.append(image)
    return images


def _nutrition_detail_images(result: dict, nutrition_results: dict,
                             nutrient_definitions: tuple[NutrientDefinition, ...]) -> list[Image.Image]:
    dishes = _ordered_dishes(result); per_page = 10; nutrients_per_page = 7; images = []
    nutrient_chunks = [nutrient_definitions[start:start + nutrients_per_page]
                       for start in range(0, len(nutrient_definitions), nutrients_per_page)]
    for nutrient_page, definitions in enumerate(nutrient_chunks, 1):
        for start in range(0, len(dishes), per_page):
            image = page_image("landscape"); draw = ImageDraw.Draw(image); margin, top = 55, 110
            draw.text((image.width//2, 42), f'{result["menu"]["name"]} 詳細營養（每人配方） {nutrient_page}/{len(nutrient_chunks)}', font=font(17), fill=GREEN, anchor="ma")
            headers = ["菜色"] + [f"{item.name}\n({item.unit})" if item.unit else item.name for item in definitions]
            widths = [330] + [(image.width - margin*2 - 330) / len(definitions)] * len(definitions)
            header_h = 82; x = margin
            for width, value in zip(widths, headers):
                draw.rectangle((x, top, x+width, top+header_h), fill=GREEN, outline=LINE, width=2)
                draw.multiline_text((x+width/2, top+header_h/2), "\n".join(wrap_text(draw, value, font(8), width-12)), font=font(8), fill="white", anchor="mm", align="center", spacing=2); x += width
            page_dishes = dishes[start:start+per_page]; row_h = (image.height-top-header_h-55)/per_page; y = top+header_h
            for dish_id, name in page_dishes:
                nutrition = nutrition_results[dish_id]; values = [name]
                for definition in definitions:
                    nutrient = nutrition.nutrients[definition.code]
                    values.append(_display_nutrition(nutrient.value) if nutrient.complete else "無")
                x = margin
                for index, (width, value) in enumerate(zip(widths, values)):
                    draw.rectangle((x, y, x+width, y+row_h), fill=PALE_GREEN if index == 0 else "white", outline=LINE, width=2)
                    lines = wrap_text(draw, value, font(8), width-12)
                    draw.multiline_text((x+width/2, y+row_h/2), "\n".join(lines), font=font(8), fill=INK, anchor="mm", align="center", spacing=2); x += width
                y += row_h
            images.append(image)
    return images


def menu_pdf(result: dict, layout: str) -> bytes:
    images = _menu_images(result, layout)
    return images_to_pdf(images, "landscape")


def menu_nutrition_pdf(result: dict, layout: str, mode: str, nutrition_results: dict,
                       nutrient_definitions: tuple[NutrientDefinition, ...]) -> bytes:
    images = []
    if mode == "detailed": images.extend(_menu_images(result, layout))
    images.extend(_menu_images(result, layout, nutrition_results))
    if mode == "detailed": images.extend(_nutrition_detail_images(result, nutrition_results, nutrient_definitions))
    return images_to_pdf(images, "landscape")


def menu_full_pdf(result: dict) -> bytes: return menu_pdf(result, "merged")
def menu_grid_pdf(result: dict) -> bytes: return menu_pdf(result, "grid")
def menu_pretty_pdf(result: dict) -> bytes: return menu_pdf(result, "pretty")
