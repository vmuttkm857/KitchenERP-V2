from datetime import date,datetime
from decimal import Decimal
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment,Border,Font,PatternFill,Side
from openpyxl.utils import get_column_letter
from app.domains.exports.safety import safe_cell_text

HEADER_FILL=PatternFill("solid",fgColor="1F4E78")
SECTION_FILL=PatternFill("solid",fgColor="D9EAF7")
THIN=Side(style="thin",color="D9E2F3")

def _value(value):
    if isinstance(value,Decimal):return int(value) if value==value.to_integral_value() else float(value)
    if isinstance(value,datetime):return value.isoformat()
    if isinstance(value,date):return value
    if isinstance(value,str):return safe_cell_text(value)
    return value

def _sheet(wb,title,headers,rows):
    ws=wb.create_sheet(title);ws.sheet_view.showGridLines=False;ws.freeze_panes="A2"
    ws.append(headers)
    for cell in ws[1]:cell.fill=HEADER_FILL;cell.font=Font(color="FFFFFF",bold=True);cell.alignment=Alignment(horizontal="center")
    for row in rows:ws.append([_value(value) for value in row])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border=Border(bottom=THIN);cell.alignment=Alignment(vertical="top",wrap_text=True)
            if isinstance(cell.value,(int,float)):cell.number_format='#,##0.########'
            elif isinstance(cell.value,(date,datetime)):cell.number_format='yyyy-mm-dd'
    ws.auto_filter.ref=ws.dimensions
    for column in range(1,len(headers)+1):
        width=max(len(str(ws.cell(row, column).value or "")) for row in range(1,min(ws.max_row,100)+1))+2
        ws.column_dimensions[get_column_letter(column)].width=min(max(width,10),36)
    return ws

def _bytes(wb):
    if "Sheet" in wb.sheetnames:del wb["Sheet"]
    stream=BytesIO();wb.save(stream);return stream.getvalue()

def kitchen_workbook(result):
    wb=Workbook()
    detail=[]
    for day in result["days"]:
        for meal in day["meals"]:
            for dish in meal["dishes"]:
                if not dish["ingredients"]:detail.append([day["menu_date"],meal["meal_type_name"],dish["dish_code"],dish["dish_name"],dish["diner_count"],None,None,None,None,None,"尚無可用配方"])
                for line in dish["ingredients"]:
                    detail.append([day["menu_date"],meal["meal_type_name"],dish["dish_code"],dish["dish_name"],dish["diner_count"],line["ingredient_code"],line["ingredient_name"],line["display_quantity"],line["display_unit"],line["notes"],"、".join(a["code"] for a in line["anomalies"])])
    _sheet(wb,"備料明細",["日期","餐別","菜色代碼","菜色","人數","食材代碼","食材","備料量","單位","備註","異常"],detail)
    _sheet(wb,"食材彙總",["食材代碼","食材","供應商","總備料量","單位","來源數","異常"],[[x["ingredient_code"],x["ingredient_name"],x["supplier_name"] or "未指定",x["display_quantity"],x["display_unit"],x["source_count"],"、".join(a["code"] for a in x["anomalies"])] for x in result["ingredient_summary"]])
    _sheet(wb,"異常",["等級","代碼","訊息","關聯資料"],[[x["severity"],x["code"],x["message"],x["related_entity_name"]] for x in result["anomalies"]])
    return _bytes(wb)

def requirements_workbook(result,title="需求量報表"):
    wb=Workbook()
    _sheet(wb,"需求彙總",["食材代碼","食材","供應商","需求量","需求單位","建議採購量","採購單位","現價","預估成本","需確認"],[[x["ingredient_code"],x["ingredient_name"],x["supplier_name"] or "未指定",x["requirement_quantity"],x["requirement_unit"],x["suggested_purchase_quantity"],x["suggested_purchase_unit"],x["current_price"],x["estimated_cost"],"是" if x["needs_review"] else "否"] for x in result["rows"]])
    daily=sorted(result["daily_rows"],key=lambda x:(x["requirement_date"],x["supplier_name"] or "未指定供應商",x["menu_name"],x["ingredient_code"],x["unit"],str(x["menu_id"]),str(x["ingredient_id"])))
    _sheet(wb,"每日採購需求",["使用日期","菜單","供應商","食材編號","食材名稱","需求量","單位"],[[x["requirement_date"],x["menu_name"],x["supplier_name"] or "未指定供應商",x["ingredient_code"],x["ingredient_name"],x["quantity"],x["unit"]] for x in daily])
    groups=[]
    by_key={x["row_key"]:x for x in result["rows"]}
    for group in result["supplier_groups"]:
        for key in group["row_keys"]:
            x=by_key[key];groups.append([group["supplier_name"],x["ingredient_code"],x["ingredient_name"],x["suggested_purchase_quantity"],x["suggested_purchase_unit"],x["estimated_cost"]])
    _sheet(wb,"供應商分組",["供應商","食材代碼","食材","建議採購量","單位","預估成本"],groups)
    _sheet(wb,"異常",["等級","代碼","訊息","關聯資料"],[[x["severity"],x["code"],x["message"],x["related_entity_name"]] for x in result["anomalies"]])
    return _bytes(wb)

def snapshot_workbook(result):
    wb=Workbook();criteria=result["criteria"] or {}
    _sheet(wb,"快照摘要",["欄位","固定值"],[["Revision",result["revision"]],["建立時間",result["created_at"]],["建立者",result["created_by_name"]],["來源菜單","、".join(x.get("menu_name","") for x in result["source_menus"])],["篩選條件",str(criteria)],["完整成本",result["total_estimated_cost"]],["已知成本",result["known_estimated_cost"]]])
    _sheet(wb,"固定品項",["食材代碼","食材","供應商","固定需求量","需求單位","系統建議","人工調整","採購單位","調整後成本","價格快照","原始成本","異常"],[[x["ingredient_code_snapshot"],x["ingredient_name_snapshot"],x["supplier_name_snapshot"] or "未指定",x["requirement_quantity"],x["requirement_unit"],x["suggested_purchase_quantity"],x["adjusted_quantity"],x["purchase_unit_snapshot"],x["adjusted_estimated_cost"],x["unit_price_snapshot"],x["estimated_cost_snapshot"],"、".join(a.get("code","") for a in x["anomaly_snapshot"])] for x in result["items"]])
    _sheet(wb,"異常快照",["等級","代碼","訊息"],[[x.get("severity"),x.get("code"),x.get("message")] for x in result["anomaly_snapshot"]])
    return _bytes(wb)

def purchase_workbook(result):
    wb=Workbook();summary=[["採購編號",result["purchase_number"]],["狀態",result["status"]],["建立時間",result["created_at"]],["建立者",result["created_by_name"]],["來源 Snapshot Revision",result["source_snapshot_revision"]],["備註",result["notes"]],["完整成本",result["total_cost"]],["已知成本",result["known_total_cost"]]]
    _sheet(wb,"採購摘要",["欄位","固定值"],summary)
    used=set()
    for index,order in enumerate(result["orders"],1):
        base=(order["supplier_name_snapshot"] or f"供應商{index}")[:25];name=base
        n=2
        while name in used:name=f"{base[:22]}_{n}";n+=1
        used.add(name)
        _sheet(wb,name,["食材代碼","食材","正式採購量","單位","包裝參考","最低量參考","單價快照","採購成本","異常"],[[x.ingredient_code_snapshot,x.ingredient_name_snapshot,x.final_purchase_quantity,x.purchase_unit_snapshot,x.package_size_snapshot,x.minimum_order_quantity_snapshot,x.unit_price_snapshot,x.purchase_cost_snapshot,"、".join(a.get("code","") for a in x.anomaly_snapshot)] for x in order["items"]])
    return _bytes(wb)
