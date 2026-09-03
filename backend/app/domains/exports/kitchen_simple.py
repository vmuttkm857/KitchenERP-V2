from __future__ import annotations

from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.pagebreak import Break
from PIL import Image, ImageDraw

from app.domains.exports.raster_pdf import GREEN, INK, LINE, MUTED, PALE_GREEN, font, images_to_pdf, page_image, wrap_text
from app.domains.exports.safety import safe_cell_text
from app.shared.domain.quantities import convert_quantity, normalize_unit

WARNING_TEXT = "⚠ 此菜有配方資料異常，請確認"
WEEKDAYS = "一二三四五六日"
THIN = Side(style="thin", color="C9D6CC")


def _warning(dish: dict) -> bool:
    return not dish["recipe_ready"] or any(a["severity"] == "error" for a in dish["anomalies"]) or any(a["severity"] == "error" for line in dish["ingredients"] for a in line["anomalies"])


def _clean(value: Decimal) -> str:
    return format(value.quantize(Decimal("0.001")).normalize(), "f")


def display_ingredient(line: dict) -> tuple[str, str]:
    quantity, unit = line.get("required_quantity"), line.get("required_unit")
    if quantity is None or not unit: return "不可計算", unit or "-"
    quantity = Decimal(str(quantity)); normalized = normalize_unit(unit)
    target = "kg" if normalized in {"g", "kg", "斤"} else ("L" if normalized in {"ml", "L"} else normalized)
    converted = convert_quantity(quantity, normalized, target)
    return (_clean(converted.quantity), target) if converted.convertible and converted.quantity is not None else (_clean(quantity), unit)


def simple_page_plan(result: dict) -> list[dict]:
    days = result["days"]
    labels_by_meal = {}
    for column in sorted(result.get("meal_type_columns", []),key=lambda item:(item["menu_meal_type_id"],item["sort_order"],str(item["id"]))):
        labels_by_meal.setdefault(column["menu_meal_type_id"],[]).append(str(column["name"]))
    meal_order, seen = [], set()
    for day in days:
        for meal in day["meals"]:
            meal_id = meal.get("meal_type_id", meal["meal_type_name"])
            if meal_id not in seen: seen.add(meal_id); meal_order.append({"id": meal_id, "name": meal["meal_type_name"],"labels":labels_by_meal.get(meal_id,[])})
    lookup = {(day["menu_date"], meal.get("meal_type_id", meal["meal_type_name"])): meal["dishes"] for day in days for meal in day["meals"]}
    plans=[]
    for start in range(0,len(days),7):
        dates=[day["menu_date"] for day in days[start:start+7]]
        meals=[]
        for meal in meal_order:
            row_count=max(len(meal["labels"]),max((len(lookup.get((day,meal["id"]),[])) for day in dates),default=0),1)
            meals.append({**meal,"row_count":row_count})
        plans.append({"dates":dates,"meals":meals,"slots":lookup})
    return plans


def _dish_lines(dish: dict) -> list[str]:
    lines = [f'{dish["dish_name"]}　{dish["diner_count"]}人']
    for ingredient in dish["ingredients"]:
        quantity, unit = display_ingredient(ingredient); lines.append(f'　• {ingredient["ingredient_name"] or "食材資料缺失"}　{quantity} {unit}')
    if _warning(dish): lines.append(WARNING_TEXT)
    return lines


def kitchen_simple_workbook(result: dict, variant: str = "single") -> bytes:
    workbook = Workbook(); del workbook["Sheet"]
    plans=simple_page_plan(result)
    for index, plan in enumerate(plans, 1):
        sheet = workbook.create_sheet("週配料表" if len(plans) == 1 else f"週配料表-{index}")
        sheet.sheet_view.showGridLines = False; sheet.page_setup.paperSize = sheet.PAPERSIZE_A4; sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
        sheet.page_margins.left = sheet.page_margins.right = .15; sheet.page_margins.top = sheet.page_margins.bottom = .2
        sheet.column_dimensions["A"].width = 11; sheet.column_dimensions["B"].width = 13
        for col in "CDEFGHI": sheet.column_dimensions[col].width = 22
        sheet.merge_cells("A1:I1"); sheet["A1"] = safe_cell_text(f'{result["menu"]["menu_name"]} 週配料表'); sheet["A1"].font = Font(size=16, bold=True, color="244B32"); sheet["A1"].alignment = Alignment(horizontal="center")
        headers = ["餐別","菜單欄位"] + [f'{day.month}/{day.day}（{WEEKDAYS[day.weekday()]}）' for day in plan["dates"]] + [""]*(7-len(plan["dates"]))
        for col, value in enumerate(headers, 1):
            cell = sheet.cell(2, col, value); cell.fill = PatternFill("solid", fgColor="356B48"); cell.font = Font(bold=True, color="FFFFFF"); cell.alignment = Alignment(horizontal="center"); cell.border = Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
        row, ends = 3, []
        base_size = 8 if len(plan["meals"]) >= 5 else 9
        if variant == "poster": base_size += 2
        for meal in plan["meals"]:
            start_row=row
            for offset in range(meal["row_count"]):
                sheet.cell(row,1,safe_cell_text(meal["name"]) if offset==0 else "")
                sheet.cell(row,2,safe_cell_text(meal["labels"][offset]) if offset<len(meal["labels"]) else "")
                max_lines=1
                for day_i in range(7):
                    dishes=plan["slots"].get((plan["dates"][day_i],meal["id"]),[]) if day_i<len(plan["dates"]) else []
                    text="\n".join(_dish_lines(dishes[offset])) if offset<len(dishes) else ""
                    max_lines=max(max_lines,text.count("\n")+1 if text else 1);sheet.cell(row,day_i+3,safe_cell_text(text))
                for col in range(1,10):
                    cell=sheet.cell(row,col);cell.font=Font(size=base_size+(2 if col==1 else 0),bold=col<=2,color="244B32" if col<=2 else INK.lstrip("#"))
                    cell.fill=PatternFill("solid",fgColor="F3F8F4" if col<=2 else "FFFFFF");cell.alignment=Alignment(horizontal="center" if col<=2 else "left",vertical="center" if col<=2 else "top",wrap_text=True);cell.border=Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
                sheet.row_dimensions[row].height=min(250,max(50,max_lines*(base_size+5)))*(1.25 if variant=="poster" else 1);row+=1
            if meal["row_count"]>1:sheet.merge_cells(start_row=start_row,start_column=1,end_row=row-1,end_column=1)
            ends.append(row-1)
        last = max(2, row-1); sheet.print_area = f"A1:I{last}"
        if variant == "single": sheet.page_setup.fitToWidth = sheet.page_setup.fitToHeight = 1; sheet.sheet_properties.pageSetUpPr.fitToPage = True
        else:
            sheet.sheet_properties.pageSetUpPr.fitToPage = False; sheet.page_setup.scale = 130; sheet.page_setup.fitToWidth = sheet.page_setup.fitToHeight = 0; sheet.page_setup.pageOrder = "overThenDown"
            sheet.col_breaks.append(Break(id=5)); sheet.row_breaks.append(Break(id=ends[(len(ends)-1)//2] if len(ends)>1 else max(3,last//2)))
    stream = BytesIO(); workbook.save(stream); return stream.getvalue()


def _draw_pdf_dish(draw: ImageDraw.ImageDraw, dish: dict, left: float, top: float, width: float, height: float) -> None:
    padding=8
    for body_size in (7,6.5,6,5.5,5,4.5,4):
        heading_font=font(body_size+1);body_font=font(body_size);warning_font=font(max(4,body_size-.5));lines=[]
        lines.extend((value,heading_font,GREEN) for value in wrap_text(draw,f'{dish["dish_name"]}  {dish["diner_count"]}人',heading_font,width-padding*2))
        for ingredient in dish["ingredients"]:
            qty,unit=display_ingredient(ingredient);text=f'• {ingredient["ingredient_name"] or "食材資料缺失"} {qty} {unit}'
            lines.extend((value,body_font,INK) for value in wrap_text(draw,text,body_font,width-padding*2))
        if _warning(dish):lines.extend((value,warning_font,"#8A5200") for value in wrap_text(draw,WARNING_TEXT,warning_font,width-padding*2))
        line_heights=[text_font.getbbox("國Ag")[3]-text_font.getbbox("國Ag")[1]+3 for _,text_font,_ in lines]
        if sum(line_heights)<=height-padding*2 or body_size==4:break
    cursor=top+padding
    for (value,text_font,color),line_height in zip(lines,line_heights):
        draw.text((left+padding,cursor),value,font=text_font,fill=color);cursor+=line_height


def kitchen_simple_pdf(result: dict) -> bytes:
    images: list[Image.Image] = []
    for plan in simple_page_plan(result):
        image = page_image("landscape"); draw = ImageDraw.Draw(image); margin, top = 55, 105
        draw.text((image.width//2, 40), f'{result["menu"]["menu_name"]} 週配料表', font=font(17), fill=GREEN, anchor="ma")
        meal_w,label_w=110,140;day_w=(image.width-margin*2-meal_w-label_w)/7;header_h=58
        headers = ["餐別","菜單欄位"]+[f'{d.month}/{d.day}（{WEEKDAYS[d.weekday()]}）' for d in plan["dates"]]+[""]*(7-len(plan["dates"])); x=margin
        for i,value in enumerate(headers):
            width=meal_w if i==0 else label_w if i==1 else day_w; draw.rectangle((x,top,x+width,top+header_h),fill=GREEN,outline=LINE,width=2); draw.text((x+width/2,top+header_h/2),value,font=font(8),fill="white",anchor="mm");x+=width
        row_h=(image.height-top-header_h-45)/max(1,len(plan["meals"])); y=top+header_h
        for meal in plan["meals"]:
            draw.rectangle((margin,y,margin+meal_w,y+row_h),fill=PALE_GREEN,outline=LINE,width=2);draw.text((margin+meal_w/2,y+row_h/2),meal["name"],font=font(9),fill=GREEN,anchor="mm")
            part=row_h/meal["row_count"]
            for offset in range(meal["row_count"]):
                row_top=y+offset*part;draw.rectangle((margin+meal_w,row_top,margin+meal_w+label_w,row_top+part),fill=PALE_GREEN,outline=LINE,width=2)
                label=meal["labels"][offset] if offset<len(meal["labels"]) else "";draw.text((margin+meal_w+label_w/2,row_top+part/2),label,font=font(8),fill=GREEN,anchor="mm")
            for day_i in range(7):
                left=margin+meal_w+label_w+day_i*day_w;dishes=plan["slots"].get((plan["dates"][day_i],meal["id"]),[]) if day_i<len(plan["dates"]) else []
                for offset in range(meal["row_count"]):
                    row_top=y+offset*part;draw.rectangle((left,row_top,left+day_w,row_top+part),fill="white",outline=LINE,width=2)
                    if offset<len(dishes):_draw_pdf_dish(draw,dishes[offset],left,row_top,day_w,part)
            y+=row_h
        images.append(image)
    return images_to_pdf(images,"landscape")
