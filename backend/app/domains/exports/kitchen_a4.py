from datetime import date
from decimal import Decimal
from io import BytesIO
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.pagebreak import Break

from app.domains.exports.safety import safe_cell_text


INVALID_SHEET_CHARACTERS = re.compile(r"[\[\]:*?/\\]")
TABLE_HEADERS = ("食材名稱", "備料量", "單位", "", "每人量", "單位")
ROWS_PER_PAGE = 24
THIN = Side(style="thin", color="808080")
MEDIUM = Side(style="medium", color="404040")
HEADER_FILL = PatternFill("solid", fgColor="D9E2F3")
DISH_FILL = PatternFill("solid", fgColor="EDEDED")


def safe_worksheet_title(value: str, used: set[str]) -> str:
    base = INVALID_SHEET_CHARACTERS.sub("_", value).strip(" '") or "作業表"
    base = base[:31]
    candidate = base
    suffix = 2
    while candidate.casefold() in used:
        marker = f"_{suffix}"
        candidate = f"{base[:31-len(marker)]}{marker}"
        suffix += 1
    used.add(candidate.casefold())
    return candidate


def paginate_dishes(dishes: list[dict], rows_per_page: int = ROWS_PER_PAGE) -> list[list[dict]]:
    """Split dish blocks without stranding their title/table header at a page bottom."""
    pages: list[list[dict]] = [[]]
    remaining = rows_per_page
    for dish in dishes:
        ingredients = dish.get("ingredients") or [None]
        offset = 0
        continuation = False
        while offset < len(ingredients):
            minimum_block_rows = 4  # dish title, table header, at least one item, spacer
            if remaining < minimum_block_rows and pages[-1]:
                pages.append([])
                remaining = rows_per_page
            available_items = max(1, remaining - 3)
            take = min(len(ingredients) - offset, available_items)
            fragment = {
                "dish": dish,
                "ingredients": ingredients[offset : offset + take],
                "continuation": continuation,
            }
            pages[-1].append(fragment)
            used_rows = take + 3
            remaining -= used_rows
            offset += take
            continuation = True
            if offset < len(ingredients):
                pages.append([])
                remaining = rows_per_page
    return pages


def _number(value):
    if value is None:
        return None
    if isinstance(value, Decimal):
        return int(value) if value == value.to_integral_value() else float(value)
    if isinstance(value, (int, float)):
        return value
    try:
        number = Decimal(str(value))
        return int(number) if number == number.to_integral_value() else float(number)
    except Exception:
        return None


def _date_text(value) -> str:
    if isinstance(value, date):
        return value.strftime("%Y/%m/%d")
    return str(value).replace("-", "/")


def _write_fragment(ws, fragment: dict, row: int) -> int:
    dish = fragment["dish"]
    dish_code = safe_cell_text(dish.get("dish_code") or "")
    dish_name = safe_cell_text(dish.get("dish_name") or "未命名菜色")
    dish_name = f"【{dish_code} {dish_name}】" if dish_code else f"【{dish_name}】"
    if fragment["continuation"]:
        dish_name = f"{dish_name}（續）"
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    cell = ws.cell(row, 1, dish_name)
    cell.font = Font(name="Microsoft JhengHei", size=16, bold=True)
    cell.fill = DISH_FILL
    cell.alignment = Alignment(vertical="center")
    cell.border = Border(bottom=THIN)
    ws.row_dimensions[row].height = 28
    row += 1
    for column, value in enumerate(TABLE_HEADERS, 1):
        cell = ws.cell(row, column, value)
        cell.font = Font(name="Microsoft JhengHei", size=13, bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border = Border(top=THIN, bottom=THIN)
    ws.row_dimensions[row].height = 24
    row += 1
    for line in fragment["ingredients"]:
        if line is None:
            values = ("尚無可用配方", None, None, None, None, None)
        else:
            values = (
                safe_cell_text(line.get("ingredient_name") or "未命名食材"),
                _number(line.get("display_quantity")),
                safe_cell_text(line.get("display_unit") or ""),
                None,
                _number(line.get("quantity_per_person")),
                safe_cell_text(line.get("recipe_unit") or ""),
            )
        for column, value in enumerate(values, 1):
            cell = ws.cell(row, column, value)
            cell.font = Font(
                name="Microsoft JhengHei",
                size=14 if column == 2 else 13,
                bold=column == 2 and value is not None,
            )
            horizontal = "right" if column in (2, 5) else "left"
            cell.alignment = Alignment(horizontal=horizontal, vertical="center", wrap_text=True)
            cell.border = Border(bottom=THIN)
            if column in (2, 5) and value is not None:
                cell.number_format = "#,##0" if float(value).is_integer() else "#,##0.######"
        ws.row_dimensions[row].height = 24
        row += 1
    ws.row_dimensions[row].height = 8
    return row + 1


def _write_page_header(
    ws,
    row: int,
    *,
    menu_name: str,
    menu_date: str,
    meal_name: str,
    page_number: int,
    page_count: int,
) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    title = ws.cell(row, 1, "廚房作業備料表")
    title.font = Font(name="Microsoft JhengHei", size=20, bold=True)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 34
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
    menu = ws.cell(row, 1, f"菜單：{menu_name}")
    menu.font = Font(name="Microsoft JhengHei", size=17, bold=True)
    menu.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    menu_date_cell = ws.cell(row, 3, f"日期：{menu_date}")
    menu_date_cell.font = Font(name="Microsoft JhengHei", size=17, bold=True)
    menu_date_cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=False)
    ws.row_dimensions[row].height = 31
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
    meal = ws.cell(row, 1, meal_name)
    meal.font = Font(name="Microsoft JhengHei", size=20, bold=True)
    meal.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    page = ws.cell(row, 3, f"第 {page_number} / {page_count} 頁")
    page.font = Font(name="Microsoft JhengHei", size=16, bold=True)
    page.alignment = Alignment(horizontal="right", vertical="center", wrap_text=False)
    ws.row_dimensions[row].height = 34
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    separator = ws.cell(row, 1)
    separator.border = Border(bottom=MEDIUM)
    ws.row_dimensions[row].height = 10
    return row + 1


def kitchen_a4_workbook(result: dict) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    used_titles: set[str] = set()
    menu_name = safe_cell_text(result["menu"]["menu_name"])
    for day in result["days"]:
        menu_date = _date_text(day["menu_date"])
        for meal in day["meals"]:
            meal_name = safe_cell_text(meal["meal_type_name"])
            title = safe_worksheet_title(f"{menu_date[5:]}_{meal_name}", used_titles)
            ws = wb.create_sheet(title)
            ws.sheet_view.showGridLines = False
            ws.sheet_view.view = "pageLayout"
            ws.column_dimensions["A"].width = 37
            ws.column_dimensions["B"].width = 17
            ws.column_dimensions["C"].width = 8
            ws.column_dimensions["D"].width = 6.5
            ws.column_dimensions["E"].width = 17
            ws.column_dimensions["F"].width = 8
            pages = paginate_dishes(meal["dishes"])
            row = 1
            for page_index, page in enumerate(pages):
                row = _write_page_header(
                    ws,
                    row,
                    menu_name=menu_name,
                    menu_date=menu_date,
                    meal_name=meal_name,
                    page_number=page_index + 1,
                    page_count=len(pages),
                )
                for fragment in page:
                    row = _write_fragment(ws, fragment, row)
                if page_index < len(pages) - 1:
                    ws.row_breaks.append(Break(id=max(1, row - 1)))
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ws.page_margins.left = 0.25
            ws.page_margins.right = 0.25
            ws.page_margins.top = 0.35
            ws.page_margins.bottom = 0.35
            ws.page_margins.header = 0
            ws.page_margins.footer = 0
            ws.print_options.horizontalCentered = True
            ws.print_area = f"A1:F{max(1, ws.max_row)}"
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()
